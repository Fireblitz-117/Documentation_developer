from openpyxl import workbook, load_workbook

def login(name, passw):
    wb = load_workbook('/Users/aaron.v/Documents/DocumentationDeveloper/Plagiarism/Python.xlsx')
    ws = wb.active
    while name and passw:
        ws.append([name, passw])
        wb.save('/Users/aaron.v/Documents/DocumentationDeveloper/Plagiarism/Python.xlsx')
        break

def logoff():
    wb = load_workbook('/Users/aaron.v/Documents/DocumentationDeveloper/Plagiarism/Python.xlsx')
    ws = wb.active
    message = 'has been logged out'
    ws.append([message])
    wb.save('/Users/aaron.v/Documents/DocumentationDeveloper/Plagiarism/Python.xlsx')